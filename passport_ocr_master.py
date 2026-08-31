"""
================================================================================
PASSPORT OCR, ARABIC TRANSLITERATION & MASTER EXCEL - ALL-IN-ONE MODULE
================================================================================
API KEY: Loaded securely from environment variables (GEMINI_API_KEY / CLIENT_GEMINI_KEY)

REQUIRES DEPENDENCIES:
pip install google-genai pandas openpyxl pillow pydantic python-dotenv
================================================================================
"""

import os
import io
import sys
import json
import re
import sqlite3
import urllib.request
import urllib.parse
import base64
from datetime import datetime, timedelta
from typing import Dict, Any, Optional

from dotenv import load_dotenv
load_dotenv()

# Force UTF-8 encoding for stdout and stderr on Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')
from pydantic import BaseModel, Field
from PIL import Image
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURATION & API KEY SETUP
# ==============================================================================
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "").strip()
CEREBRAS_API_KEY = os.getenv("CEREBRAS_API_KEY", "").strip()
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "").strip()
DB_FILE = os.path.join(os.path.dirname(__file__), "passports.db")
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Master_Passports.xlsx")

def get_gemini_api_keys(custom_key: Optional[str] = None) -> list:
    keys = []
    client_key = os.getenv("CLIENT_GEMINI_KEY", "").strip()
    if client_key:
        keys.extend([k.strip() for k in client_key.split(",") if k.strip()])
    if custom_key:
        keys.extend([k.strip() for k in custom_key.split(",") if k.strip()])
    env_keys = os.getenv("GEMINI_API_KEY", "")
    if env_keys:
        keys.extend([k.strip() for k in env_keys.split(",") if k.strip()])
    env_key2 = os.getenv("GEMINI_API_KEY_2", "")
    if env_key2:
        keys.extend([k.strip() for k in env_key2.split(",") if k.strip()])
    unique_keys = []
    for k in keys:
        if k not in unique_keys:
            unique_keys.append(k)
    return unique_keys if unique_keys else [GEMINI_API_KEY]


# ==============================================================================
# PYDANTIC SCHEMAS
# ==============================================================================
class PassportSchema(BaseModel):
    first_name: str = Field(..., description="Given name(s) in English capital letters or 'N/A' if blank/single-name passport")
    last_name: str = Field(..., description="Surname/Last name in English capital letters or 'N/A' if blank/single-name passport")
    father_name: Optional[str] = Field(None, description="Father's name field exactly as written on passport (e.g. 'NAZIR, MUHAMMAD')")
    passport_number: str = Field(..., description="Unique alphanumeric passport identification number")
    nationality: str = Field(..., description="The 3-letter ISO country code or nationality name")
    date_of_birth: str = Field(..., description="Date of birth in YYYY-MM-DD format")
    date_of_issue: str = Field(..., description="Passport issuance date in YYYY-MM-DD format")
    date_of_expiry: str = Field(..., description="Passport expiry date in YYYY-MM-DD format")

class ArabicTranslationSchema(BaseModel):
    first_name_ar: str = Field(..., description="First name phonetically transliterated into Arabic script")
    last_name_ar: str = Field(..., description="Surname phonetically transliterated into Arabic script")
    nationality_ar: str = Field(..., description="Country name translated into standard Arabic text")

class TicketSchema(BaseModel):
    departure_date: str = Field(..., description="Outbound flight departure date in YYYY-MM-DD format (e.g. '2026-11-03')")
    return_date: Optional[str] = Field(None, description="Return flight date in YYYY-MM-DD format if round-trip (e.g. '2026-11-17'), or 'N/A'")
    airline_name: Optional[str] = Field(None, description="Airline carrier name e.g. 'Saudi Arabian Airlines', 'PIA', 'AirBlue'")
    flight_numbers: Optional[str] = Field(None, description="Flight codes/numbers e.g. 'SV 735 / SV 734'")
    origin_city: Optional[str] = Field(None, description="Departure city in Pakistan e.g. 'Lahore', 'Karachi', 'Islamabad'")
    destination_city: Optional[str] = Field(None, description="Arrival city in Saudi Arabia e.g. 'Jeddah', 'Madinah'")
    arrival_airport: Optional[str] = Field(None, description="Destination/arrival airport or city e.g. 'Jeddah', 'JED', 'Madinah', 'MED'")

# ==============================================================================
# DATABASE MANAGEMENT (SQLite)
# ==============================================================================
def initialize_db():
    """Creates the passports table if it doesn't already exist."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS passport_records (
            passport_number TEXT PRIMARY KEY,
            first_name TEXT,
            last_name TEXT,
            nationality TEXT,
            date_of_birth TEXT,
            date_of_issue TEXT,
            date_of_expiry TEXT,
            first_name_ar TEXT,
            last_name_ar TEXT,
            nationality_ar TEXT,
            customer_phone TEXT,
            request_id TEXT,
            status TEXT DEFAULT 'Pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    
    # Auto-migrate table if columns are missing
    cursor.execute("PRAGMA table_info(passport_records)")
    columns = [row[1] for row in cursor.fetchall()]
    if 'date_of_issue' not in columns:
        cursor.execute("ALTER TABLE passport_records ADD COLUMN date_of_issue TEXT")
    if 'customer_phone' not in columns:
        cursor.execute("ALTER TABLE passport_records ADD COLUMN customer_phone TEXT")
    if 'request_id' not in columns:
        cursor.execute("ALTER TABLE passport_records ADD COLUMN request_id TEXT")
    conn.commit()
    conn.close()

initialize_db()

def save_pending_record(data: Dict[str, Any], phone: str = "", request_id: str = "") -> Dict[str, Any]:
    """Saves OCR extracted data in 'Pending' state."""
    pno = str(data.get('passport_number') or data.get('passportNumber') or 'N/A').upper()
    fn = data.get('first_name') or data.get('firstName') or ''
    ln = data.get('last_name') or data.get('lastName') or ''
    nat = data.get('nationality') or 'Pakistani'
    dob = data.get('date_of_birth') or data.get('dob') or 'N/A'
    iss = data.get('date_of_issue') or data.get('issueDate') or 'N/A'
    exp = data.get('date_of_expiry') or data.get('expiryDate') or 'N/A'

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR REPLACE INTO passport_records 
        (passport_number, first_name, last_name, nationality, date_of_birth, date_of_issue, date_of_expiry, customer_phone, request_id, status, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending', CURRENT_TIMESTAMP)
    ''', (
        pno, fn, ln,
        nat, dob, iss, exp,
        phone or data.get('customer_phone', ''), request_id or data.get('request_id', '')
    ))
    conn.commit()
    conn.close()
    return get_record(pno)

def update_confirmed_record(passport_number: str, english_data: Dict[str, Any], arabic_data: Dict[str, Any], phone: str = "", request_id: str = "") -> Dict[str, Any]:
    """Updates record with confirmed English and Arabic data, setting status to 'Confirmed'."""
    pno = str(passport_number or english_data.get('passport_number') or english_data.get('passportNumber') or 'N/A').upper()
    fn = english_data.get('first_name') or english_data.get('firstName') or ''
    ln = english_data.get('last_name') or english_data.get('lastName') or ''
    nat = english_data.get('nationality') or 'Pakistani'
    dob = english_data.get('date_of_birth') or english_data.get('dob') or 'N/A'
    iss = english_data.get('date_of_issue') or english_data.get('issueDate') or 'N/A'
    exp = english_data.get('date_of_expiry') or english_data.get('expiryDate') or 'N/A'
    fn_ar = arabic_data.get('first_name_ar') or arabic_data.get('firstNameAr') or translate_single_field_to_arabic(fn)
    ln_ar = arabic_data.get('last_name_ar') or arabic_data.get('lastNameAr') or translate_single_field_to_arabic(ln)
    nat_ar = arabic_data.get('nationality_ar') or arabic_data.get('nationalityAr') or nat

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE passport_records 
        SET first_name = ?, last_name = ?, nationality = ?, date_of_birth = ?, date_of_issue = ?, date_of_expiry = ?,
            first_name_ar = ?, last_name_ar = ?, nationality_ar = ?,
            customer_phone = COALESCE(NULLIF(?, ''), customer_phone),
            request_id = COALESCE(NULLIF(?, ''), request_id),
            status = 'Confirmed', updated_at = CURRENT_TIMESTAMP
        WHERE passport_number = ?
    ''', (
        fn, ln, nat,
        dob, iss, exp,
        fn_ar, ln_ar, nat_ar,
        phone, request_id,
        pno
    ))
    conn.commit()
    conn.close()
    return get_record(pno)

def get_record(passport_number: str) -> Optional[Dict[str, Any]]:
    """Retrieves a record by passport number."""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM passport_records WHERE passport_number = ?", (passport_number,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None

# ==============================================================================
# ARABIC PHONETIC TRANSLITERATION ENGINE & DICTIONARY
# ==============================================================================
ARABIC_NAME_DICT = {
    "MUHAMMAD": "محمد", "MOHAMMAD": "محمد", "MOHAMED": "محمد", "MD": "محمد", "MOHD": "محمد", "MOHMED": "محمد",
    "AHMED": "أحمد", "AHMAD": "أحمد",
    "ALI": "علي",
    "KHAN": "خان",
    "MALIK": "مالك",
    "AKRAM": "أكرم",
    "AKHTAR": "أختر",
    "RAMZAN": "رمضان", "RAMADAN": "رمضان",
    "SALEEM": "سليم", "SALIM": "سليم",
    "GHULAM": "غلام",
    "MURTAZA": "مرتضى", "MURTZA": "مرتضى",
    "RABIA": "رابعة", "RABIYA": "رابعة", "RABYA": "رابعة",
    "ANWAR": "أنور", "ANWER": "أنور",
    "HUSSAIN": "حسين", "HUSAIN": "حسين", "HOSSAIN": "حسين",
    "HASSAN": "حسن", "HASAN": "حسن",
    "USMAN": "عثمان", "OTHMAN": "عثمان", "OSMAN": "عثمان",
    "UMAR": "عمر", "OMAR": "عمر", "UMER": "عمر",
    "ABUBAKAR": "أبو بكر", "ABU BAKR": "أبو بكر", "ABUBAKR": "أبو بكر",
    "FATIMA": "فاطمة", "FATIMAH": "فاطمة",
    "AYESHA": "عائشة", "AISHA": "عائشة",
    "ZAINAB": "زينب",
    "KHADIJA": "خديجة", "KHADIJAH": "خديجة",
    "MARIAM": "مريم", "MARYAM": "مريم",
    "SADIA": "سعدية",
    "SAIMA": "صائمة",
    "FARHANA": "فرحانة",
    "AMNA": "آمنة",
    "BUSHRA": "بشرى",
    "SHAHID": "شاهد",
    "TARIQ": "طارق",
    "RASHID": "راشد",
    "KHALID": "خالد",
    "ASIF": "عاصف",
    "ARIF": "عارف",
    "IMRAN": "عمران",
    "IRFAN": "عرفان",
    "BILAL": "بلال",
    "HAMZA": "حمزة",
    "ZULFIQAR": "ذو الفقار",
    "ABDUL": "عبد ال", "ABD": "عبد", "ABDULLAH": "عبد الله", "ABDUR": "عبد ال",
    "REHMAN": "رحمن", "RAHMAN": "رحمن", "RAHEEM": "رحيم", "RAHIM": "رحيم", "RASHEED": "رشيد",
    "WAQAS": "وقاص",
    "FAISAL": "فيصل",
    "NAEEM": "نعيم",
    "NADEEM": "نديم",
    "WAHEED": "وحيد",
    "SAEED": "سعيد",
    "JAVAID": "جاويد", "JAVED": "جاويد",
    "IQBAL": "إقبال",
    "ASGHAR": "أصغر",
    "SHAH": "شاه",
    "SYED": "سيد",
    "CHAUDHRY": "تشودري", "CHAUDHARY": "تشودري", "CH": "تشودري",
    "BHATTI": "بهتي",
    "BUTT": "بت",
    "SHEIKH": "شيخ",
    "QURESHI": "قريشي",
    "SIDDIQUI": "صديقي", "SIDDIQI": "صديقي",
    "BAIG": "بيك", "BEG": "بيك",
    "MIAN": "ميان",
    "BIBI": "بي بي",
    "BEGUM": "بيكم",
    "MAQBOOL": "مقبول",
    "MANZOOR": "منظور",
    "MUSTAFA": "مصطفى",
    "ZIA": "ضياء",
    "IJAZ": "إعجاز",
    "AYOUB": "أيوب", "AYUB": "أيوب",
    "YASIR": "ياسر",
    "YASMEEN": "ياسمين",
    "NASREEN": "نسرين",
    "PARVEEN": "برفين",
    "TAHIRA": "طاهرة",
    "SAJID": "ساجد",
    "MAJID": "ماجد",
    "ABID": "عابد",
    "ZAHID": "زاهد",
    "WAJID": "واجد",
    "SHABBIR": "شبير",
    "BASHIR": "بشير",
    "NAZIR": "نظير",
    "MUNAWAR": "منور",
    "RIAZ": "رياض",
    "LIAQAT": "لياقت",
    "SHAFEEQ": "شفيق",
    "RAFIQ": "رفيق",
    "SULTAN": "سلطان",
    "FAROOQ": "فاروق",
    "SHOUKAT": "شوكت",
    "LIAQUAT": "لياقت",
    "ASAD": "أسد",
    "SARFRAZ": "سرفراز",
    "ZAFAR": "ظفر",
    "AZHAR": "أزهر",
    "SHAHBAZ": "شهباز",
    "KAMRAN": "كامران",
    "ADNAN": "عدنان",
    "NOMAN": "نعمان", "NUMAN": "نعمان",
    "ARSLAN": "أرسلان",
    "REHAN": "ريحان",
    "SOHAIL": "سهيل", "SUHAIL": "سهيل",
    "TANVEER": "تنوير", "TANVIR": "تنوير",
    "JAVERIA": "جويرية",
    "HINA": "حنا",
    "KIRAN": "كيران",
    "SANA": "ثناء",
    "IQRA": "إقرأ",
    "SIDRA": "سدرة",
    "NIMRA": "نمرة",
    "MEHREEN": "مهرين",
    "ALAM": "عالم",
    "DIN": "الدين", "DEEN": "الدين",
    "ULLAH": "الله",
    "UR": "ال",
}

def is_arabic_text(text: str) -> bool:
    if not text or not isinstance(text, str):
        return False
    return any('\u0600' <= char <= '\u06FF' or '\u0750' <= char <= '\u077F' or '\u08A0' <= char <= '\u08FF' for char in text)

def transliterate_phonetic_word(word: str) -> str:
    """Fallback character-by-character transliteration of unknown English words into Arabic script."""
    w = word.upper().strip()
    if not w:
        return ""
    if is_arabic_text(w):
        return w
    if w in ARABIC_NAME_DICT:
        return ARABIC_NAME_DICT[w]

    replacements = [
        ("KH", "خ"), ("GH", "غ"), ("SH", "ش"), ("CH", "تش"),
        ("TH", "ث"), ("DH", "ذ"), ("ZH", "ژ"), ("PH", "ف"),
        ("OU", "و"), ("OO", "و"), ("EE", "ي"), ("AI", "اي"), ("EI", "اي"),
        ("AU", "او"), ("AL-", "ال"), ("AL", "ال"),
    ]
    for eng, ar in replacements:
        w = w.replace(eng, ar)

    char_map = {
        'A': 'ا', 'B': 'ب', 'C': 'ك', 'D': 'د', 'E': 'ي',
        'F': 'ف', 'G': 'ج', 'H': 'ه', 'I': 'ي', 'J': 'ج',
        'K': 'ك', 'L': 'ل', 'M': 'م', 'N': 'ن', 'O': 'و',
        'P': 'ب', 'Q': 'ق', 'R': 'ر', 'S': 'س', 'T': 'ت',
        'U': 'و', 'V': 'ف', 'W': 'و', 'X': 'كس', 'Y': 'ي', 'Z': 'ز'
    }
    res = []
    for c in w:
        if c in char_map:
            res.append(char_map[c])
        else:
            res.append(c)
    return "".join(res)

def translate_single_field_to_arabic(text: str) -> str:
    """Translates a full name or single field into authentic Arabic script."""
    if not text or not str(text).strip() or str(text).strip().upper() in ['N/A', 'NOT DETECTED', 'NONE']:
        return "N/A"
    clean = str(text).strip()
    if is_arabic_text(clean):
        return clean

    tokens = clean.replace(',', ' ').replace('-', ' ').split()
    translated_tokens = []
    for tok in tokens:
        up = tok.upper().strip()
        if not up:
            continue
        if up in ARABIC_NAME_DICT:
            translated_tokens.append(ARABIC_NAME_DICT[up])
        else:
            translated_tokens.append(transliterate_phonetic_word(up))

    out = " ".join(translated_tokens).strip()
    out = out.replace("عبد ال ", "عبد ال")
    return out or clean

# ==============================================================================
# EXCEL EXPORTER SERVICE (With Blank Lines Between User Requests)
# ==============================================================================
def export_confirmed_passports_to_excel(request_id: Optional[str] = None, direct_passengers: Optional[List[Dict[str, Any]]] = None) -> str:
    """Exports 'Confirmed' passports into styled Master_Passports.xlsx. If direct_passengers or request_id/phone specified, exports strictly for that booking window."""
    conn = sqlite3.connect(DB_FILE)
    
    if direct_passengers and isinstance(direct_passengers, list) and len(direct_passengers) > 0:
        rows = []
        clean_req = str(request_id or "ORDER").strip()
        cursor = conn.cursor()
        for p in direct_passengers:
            fn = p.get('firstName') or p.get('first_name') or ''
            ln = p.get('lastName') or p.get('last_name') or ''
            fn_ar = p.get('firstNameAr') or p.get('first_name_ar') or ''
            ln_ar = p.get('lastNameAr') or p.get('last_name_ar') or ''

            if not is_arabic_text(fn_ar):
                fn_ar = translate_single_field_to_arabic(fn)
            if not is_arabic_text(ln_ar):
                ln_ar = translate_single_field_to_arabic(ln)

            pno = str(p.get('passportNumber') or p.get('passport_number') or 'N/A').upper()
            nat = p.get('nationality') or 'Pakistani'
            dob = p.get('dob') or p.get('date_of_birth') or 'N/A'
            iss = p.get('issueDate') or p.get('date_of_issue') or 'N/A'
            exp = p.get('expiryDate') or p.get('date_of_expiry') or 'N/A'
            
            # Sync to SQLite passport_records
            cursor.execute('''
                INSERT OR REPLACE INTO passport_records 
                (passport_number, first_name, last_name, nationality, date_of_birth, date_of_issue, date_of_expiry, first_name_ar, last_name_ar, nationality_ar, customer_phone, request_id, status, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Confirmed', CURRENT_TIMESTAMP)
            ''', (pno, fn, ln, nat, dob, iss, exp, fn_ar, ln_ar, nat, clean_req, clean_req))

            rows.append({
                "Customer Phone": clean_req,
                "Request ID": clean_req,
                "Passport Number / رقم الجواز": pno,
                "First Name (English)": fn,
                "Last Name (English)": ln,
                "First Name (Arabic) / الاسم الأول": fn_ar,
                "Last Name (Arabic) / اسم العائلة": ln_ar,
                "Nationality / الجنسية": nat,
                "Date of Birth / تاريخ الميلاد": dob,
                "Issue Date / تاريخ الإصدار": iss,
                "Expiry Date / تاريخ الانتهاء": exp,
                "Status": "Confirmed",
                "Confirmed Date": datetime.now().strftime('%Y-%m-%d')
            })
        conn.commit()
        df = pd.DataFrame(rows)
    elif request_id and str(request_id).strip():
        clean_req = str(request_id).strip()
        df = pd.read_sql_query("""
            SELECT 
                customer_phone AS "Customer Phone",
                request_id AS "Request ID",
                passport_number AS "Passport Number / رقم الجواز",
                first_name AS "First Name (English)",
                last_name AS "Last Name (English)",
                first_name_ar AS "First Name (Arabic) / الاسم الأول",
                last_name_ar AS "Last Name (Arabic) / اسم العائلة",
                nationality AS "Nationality / الجنسية",
                date_of_birth AS "Date of Birth / تاريخ الميلاد",
                date_of_issue AS "Issue Date / تاريخ الإصدار",
                date_of_expiry AS "Expiry Date / تاريخ الانتهاء",
                status AS "Status",
                updated_at AS "Confirmed Date"
            FROM passport_records 
            WHERE status = 'Confirmed' AND (request_id = ? OR customer_phone LIKE ?)
            ORDER BY updated_at ASC, request_id ASC
        """, conn, params=(clean_req, f"%{clean_req}%"))
    else:
        df = pd.read_sql_query("""
            SELECT 
                customer_phone AS "Customer Phone",
                request_id AS "Request ID",
                passport_number AS "Passport Number / رقم الجواز",
                first_name AS "First Name (English)",
                last_name AS "Last Name (English)",
                first_name_ar AS "First Name (Arabic) / الاسم الأول",
                last_name_ar AS "Last Name (Arabic) / اسم العائلة",
                nationality AS "Nationality / الجنسية",
                date_of_birth AS "Date of Birth / تاريخ الميلاد",
                date_of_issue AS "Issue Date / تاريخ الإصدار",
                date_of_expiry AS "Expiry Date / تاريخ الانتهاء",
                status AS "Status",
                updated_at AS "Confirmed Date"
            FROM passport_records 
            WHERE status = 'Confirmed'
            ORDER BY updated_at ASC, request_id ASC
        """, conn)
    conn.close()

    display_cols = [
        "Customer Phone", "Passport Number / رقم الجواز", "First Name (English)", "Last Name (English)",
        "First Name (Arabic) / الاسم الأول", "Last Name (Arabic) / اسم العائلة",
        "Nationality / الجنسية", "Date of Birth / تاريخ الميلاد", "Issue Date / تاريخ الإصدار", "Expiry Date / تاريخ الانتهاء",
        "Status", "Confirmed Date"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Passports'

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    # Write Headers
    for col_idx, col_name in enumerate(display_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    current_user_group = None
    row_write_idx = 2

    if not df.empty:
        for _, row_data in df.iterrows():
            user_group = row_data.get('Request ID') or row_data.get('Customer Phone') or 'default'
            
            # Insert a blank line between different user requests
            if current_user_group is not None and user_group != current_user_group:
                row_write_idx += 1
            
            current_user_group = user_group

            for col_idx, col_name in enumerate(display_cols, start=1):
                val = row_data.get(col_name)
                cell = ws.cell(row=row_write_idx, column=col_idx, value="" if pd.isna(val) else str(val))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_write_idx += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    ws.row_dimensions[1].height = 28
    wb.save(EXCEL_FILE)
    return EXCEL_FILE

# ==============================================================================
# OCR & TRANSLATION ENGINE
# ==============================================================================
def compress_image_if_needed(image_bytes: bytes, max_dim: int = 1200) -> tuple[bytes, str]:
    """Resizes photos for fast processing and optimal token cost."""
    try:
        img = Image.open(io.BytesIO(image_bytes))
        if img.mode != 'RGB':
            img = img.convert('RGB')
        width, height = img.size
        if width > max_dim or height > max_dim:
            ratio = max_dim / float(max(width, height))
            img = img.resize((int(width * ratio), int(height * ratio)), Image.Resampling.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=85)
        return buf.getvalue(), "image/jpeg"
    except Exception:
        return image_bytes, "image/jpeg"


def apply_father_name_rule(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    SINGLE-NAME PASSPORT RULE:
    If a passport bearer has only one name (e.g. Given Name is blank, or Surname is blank/N/A):
    - Combine the single name with Father's Name.
    - If Father's Name contains a comma ',' (e.g. "NAZIR, MUHAMMAD"):
      It represents [LAST_NAME], [FIRST_NAME]. Reverse around comma to get "MUHAMMAD NAZIR".
      Example: ABDULLAH + NAZIR, MUHAMMAD -> ABDULLAH MUHAMMAD NAZIR.
    - If Father's Name does NOT contain a comma (e.g. "NAZIR MUHAMMAD"):
      Read it directly as "NAZIR MUHAMMAD".
      Example: ABDULLAH + NAZIR MUHAMMAD -> ABDULLAH NAZIR MUHAMMAD.
    """
    first = (data.get('first_name') or '').strip()
    last = (data.get('last_name') or '').strip()
    father = (data.get('father_name') or '').strip()

    is_first_missing = not first or first.upper() in ['N/A', 'NONE', '-', 'NOT DETECTED']
    is_last_missing = not last or last.upper() in ['N/A', 'NONE', '-', 'NOT DETECTED']

    if is_first_missing or is_last_missing:
        single_name = last if is_first_missing else first

        if father:
            if ',' in father:
                parts = [p.strip() for p in father.split(',') if p.strip()]
                father_parsed = " ".join(reversed(parts))
            else:
                father_parsed = father

            data['first_name'] = single_name
            data['last_name'] = father_parsed
        else:
            data['first_name'] = single_name
            data['last_name'] = 'N/A'

    return data

def run_gemini_vision_ocr(image_bytes: bytes, api_key: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """Primary Vision OCR using Google Gemini models."""
    keys = get_gemini_api_keys(api_key)
    try:
        from google import genai
        from google.genai import types
        compressed_bytes, mime_type = compress_image_if_needed(image_bytes)

        prompt = """
        Extract all passport data accurately from this document photo.
        Carefully read visual text fields (Surname, Given Names, Father Name, Passport Number, Nationality, DOB, Date of Issue, Expiry Date).
        Cross-reference with Machine Readable Zone (MRZ) characters at the bottom.
        If Father Name is present on passport, extract it into father_name (e.g. "NAZIR, MUHAMMAD" or "NAZIR MUHAMMAD").
        Ensure date fields (date_of_birth, date_of_issue, date_of_expiry) are strictly formatted as YYYY-MM-DD.
        """

        candidate_models = ['gemini-flash-lite-latest', 'gemini-flash-latest', 'gemini-pro-latest', 'gemini-3-flash-preview']
        response = None
        import time

        for k in keys:
            key_exhausted = False
            client = genai.Client(api_key=k)
            for m in candidate_models:
                if key_exhausted:
                    break
                for attempt in range(2):
                    try:
                        response = client.models.generate_content(
                            model=m,
                            contents=[types.Part.from_bytes(data=compressed_bytes, mime_type=mime_type), prompt],
                            config=types.GenerateContentConfig(
                                response_mime_type="application/json",
                                response_schema=PassportSchema,
                                temperature=0.0
                            )
                        )
                        if response and response.text:
                            break
                    except Exception as err:
                        err_msg = str(err)
                        sys.stderr.write(f"Gemini Model {m} error: {err_msg}\n")
                        if '429' in err_msg or 'RESOURCE_EXHAUSTED' in err_msg:
                            key_exhausted = True
                            break
                        break
                if response and response.text:
                    break
            if response and response.text:
                break

        if response and response.text:
            data = PassportSchema.model_validate_json(response.text).model_dump()
            return apply_father_name_rule(data)
    except Exception as e:
        sys.stderr.write(f"Gemini OCR exception: {e}\n")

    return None

def is_dummy_placeholder(data: Dict[str, Any]) -> bool:
    """Detects if Vision AI returned dummy/mock placeholder data (e.g. John Doe / 123456789)."""
    fn = str(data.get('first_name', '')).upper().strip()
    ln = str(data.get('last_name', '')).upper().strip()
    pno = str(data.get('passport_number', '')).upper().strip()
    nat = str(data.get('nationality', '')).upper().strip()
    
    if fn in ['JOHN', 'FIRSTNAME', 'SAMPLE'] and ln in ['DOE', 'LASTNAME', 'SPECIMEN']:
        return True
    if pno in ['123456789', 'AT0000000', '000000000', 'XXXXXXXXX']:
        return True
    if (fn == 'JOHN' or ln == 'DOE') and (pno == '123456789' or nat in ['AMERICAN', 'USA']):
        return True
    return False

def run_openrouter_vision_ocr(image_bytes: bytes) -> Optional[Dict[str, Any]]:
    """Fallback Vision OCR using OpenRouter free vision models."""
    key = OPENROUTER_API_KEY or os.getenv("OPENROUTER_API_KEY", "").strip()
    if not key:
        return None
    try:
        compressed_bytes, mime_type = compress_image_if_needed(image_bytes)
        b64_data = base64.b64encode(compressed_bytes).decode("utf-8")
        data_url = f"data:{mime_type};base64,{b64_data}"

        models = [
            'openrouter/free',
            'meta-llama/llama-3.2-11b-vision-instruct:free',
            'qwen/qwen-2-vl-72b-instruct:free',
            'google/gemma-4-31b-it:free',
            'nvidia/nemotron-nano-12b-v2-vl:free'
        ]
        prompt = (
            "Extract all passport data accurately from this document photo.\n"
            "Return ONLY a JSON object with keys: first_name, last_name, father_name, passport_number, nationality, date_of_birth, date_of_issue, date_of_expiry.\n"
            "Ensure date fields are strictly formatted as YYYY-MM-DD."
        )

        for m in models:
            try:
                payload = {
                    "model": m,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                {"type": "image_url", "image_url": {"url": data_url}}
                            ]
                        }
                    ],
                    "temperature": 0.0
                }
                req = urllib.request.Request(
                    "https://openrouter.ai/api/v1/chat/completions",
                    headers={
                        "Authorization": f"Bearer {key}",
                        "Content-Type": "application/json",
                        "HTTP-Referer": "https://localhost",
                        "User-Agent": "Mozilla/5.0"
                    },
                    data=json.dumps(payload).encode("utf-8")
                )
                with urllib.request.urlopen(req, timeout=15) as res:
                    body = json.loads(res.read().decode("utf-8"))
                    content = body["choices"][0]["message"]["content"]
                    if "```json" in content:
                        content = content.split("```json")[1].split("```")[0].strip()
                    elif "```" in content:
                        content = content.split("```")[1].split("```")[0].strip()
                    raw_dict = json.loads(content)
                    data = PassportSchema.model_validate(raw_dict).model_dump()
                    if is_dummy_placeholder(data):
                        sys.stderr.write(f"[Fallback Vision OCR] OpenRouter ({m}) returned mock placeholder data, skipping...\n")
                        continue
                    sys.stderr.write(f"[Fallback Vision OCR] OpenRouter ({m}) succeeded.\n")
                    return apply_father_name_rule(data)
            except Exception as e:
                sys.stderr.write(f"[Fallback Vision OCR] OpenRouter ({m}) error: {e}\n")
                continue
    except Exception as exc:
        sys.stderr.write(f"[Fallback Vision OCR] OpenRouter exception: {exc}\n")

    return None

def run_groq_vision_ocr(image_bytes: bytes) -> Optional[Dict[str, Any]]:
    """Fallback Vision OCR using Groq Vision models."""
    key = GROQ_API_KEY or os.getenv("GROQ_API_KEY", "").strip()
    if not key:
        return None
    try:
        compressed_bytes, mime_type = compress_image_if_needed(image_bytes)
        b64_data = base64.b64encode(compressed_bytes).decode("utf-8")
        data_url = f"data:{mime_type};base64,{b64_data}"

        models = ['llama-3.2-90b-vision-preview', 'llama-3.2-11b-vision-instruct']
        prompt = (
            "Extract all passport data accurately from this document photo.\n"
            "Return ONLY a JSON object with keys: first_name, last_name, father_name, passport_number, nationality, date_of_birth, date_of_issue, date_of_expiry.\n"
            "Ensure date fields are strictly formatted as YYYY-MM-DD."
        )

        for m in models:
            try:
                payload = {
                    "model": m,
                    "messages": [
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                {"type": "image_url", "image_url": {"url": data_url}}
                            ]
                        }
                    ],
                    "temperature": 0.0
                }
                req = urllib.request.Request(
                    "https://api.groq.com/openai/v1/chat/completions",
                    headers={
                        "Authorization": f"Bearer {key}",
                        "Content-Type": "application/json",
                        "User-Agent": "Mozilla/5.0"
                    },
                    data=json.dumps(payload).encode("utf-8")
                )
                with urllib.request.urlopen(req, timeout=15) as res:
                    body = json.loads(res.read().decode("utf-8"))
                    content = body["choices"][0]["message"]["content"]
                    if "```json" in content:
                        content = content.split("```json")[1].split("```")[0].strip()
                    elif "```" in content:
                        content = content.split("```")[1].split("```")[0].strip()
                    raw_dict = json.loads(content)
                    data = PassportSchema.model_validate(raw_dict).model_dump()
                    if is_dummy_placeholder(data):
                        sys.stderr.write(f"[Fallback Vision OCR] Groq Vision ({m}) returned mock placeholder data, skipping...\n")
                        continue
                    sys.stderr.write(f"[Fallback Vision OCR] Groq Vision ({m}) succeeded.\n")
                    return apply_father_name_rule(data)
            except Exception as e:
                sys.stderr.write(f"[Fallback Vision OCR] Groq Vision ({m}) error: {e}\n")
                continue
    except Exception as exc:
        sys.stderr.write(f"[Fallback Vision OCR] Groq Vision exception: {exc}\n")

    return None

def run_passport_ocr(image_bytes: bytes, api_key: Optional[str] = None) -> Dict[str, Any]:
    """Performs Passport OCR with Multi-Provider Fallbacks (Gemini -> OpenRouter -> Groq)."""
    # Tier 1: Gemini Vision
    res = run_gemini_vision_ocr(image_bytes, api_key)
    if res:
        return res

    # Tier 2: OpenRouter Vision
    sys.stderr.write("[OCR Engine] Gemini Vision quota reached. Triggering OpenRouter Vision fallback...\n")
    res = run_openrouter_vision_ocr(image_bytes)
    if res:
        return res

    # Tier 3: Groq Vision
    sys.stderr.write("[OCR Engine] OpenRouter Vision fallback trigger -> Groq Vision...\n")
    res = run_groq_vision_ocr(image_bytes)
    if res:
        return res

    return None

def validate_passport_validity(expiry_date_str: str) -> tuple[bool, str]:
    """
    Checks if passport expiry date is valid for at least 6 months (180 days) from today.
    Returns (is_valid, error_whatsapp_message)
    """
    if not expiry_date_str or expiry_date_str.strip().upper() in ['N/A', 'NONE', 'NOT DETECTED']:
        return True, ""
    
    try:
        exp_date = None
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y']:
            try:
                exp_date = datetime.strptime(expiry_date_str.strip(), fmt)
                break
            except ValueError:
                continue
        
        if not exp_date:
            return True, ""
        
        today = datetime.now()
        six_months_ahead = today + timedelta(days=180)
        
        if exp_date < six_months_ahead:
            formatted_exp = exp_date.strftime('%d-%b-%Y')
            days_left = (exp_date - today).days
            
            if days_left <= 0:
                msg = (
                    "❌ *Passport Expired!*\n\n"
                    f"Your passport expired on *{formatted_exp}*.\n\n"
                    "⚠️ *Rule:* Your passport must be valid for *at least 6 months* to process a Hajj or Umrah visa.\n\n"
                    "Please upload a valid, renewed passport photo to proceed."
                )
            else:
                msg = (
                    "⚠️ *Passport Validity Error (Less than 6 Months)*\n\n"
                    f"• *Expiry Date:* {formatted_exp}\n"
                    f"• *Remaining Validity:* {days_left} days\n\n"
                    "⛔ *Rule:* Your passport must be valid for *at least 6 months (180 days)* from today to process a Hajj/Umrah visa.\n\n"
                    "Please upload a renewed passport with more than 6 months validity to proceed."
                )
            return False, msg
        
        return True, ""
    except Exception:
        return True, ""

def run_gemini_arabic_translation(english_data: Dict[str, Any], api_key: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """Primary Arabic Translation using Gemini."""
    keys = get_gemini_api_keys(api_key)
    if not keys:
        return None
    try:
        from google import genai
        from google.genai import types

        fn = english_data.get('first_name') or english_data.get('firstName') or ''
        ln = english_data.get('last_name') or english_data.get('lastName') or ''
        nat = english_data.get('nationality') or 'Pakistani'

        prompt = f"""
        Phonetically transliterate the person's English first_name and last_name into official Arabic script.
        Do NOT translate the country/nationality name — set nationality_ar exactly equal to Nationality as provided.
        First Name: {fn}
        Last Name: {ln}
        Nationality: {nat}
        """

        candidate_models = ['gemini-flash-latest', 'gemini-flash-lite-latest']
        response = None
        for k in keys:
            key_exhausted = False
            client = genai.Client(api_key=k)
            for m in candidate_models:
                if key_exhausted:
                    break
                try:
                    response = client.models.generate_content(
                        model=m,
                        contents=prompt,
                        config=types.GenerateContentConfig(
                            response_mime_type="application/json",
                            response_schema=ArabicTranslationSchema,
                            temperature=0.1
                        )
                    )
                    if response and response.text:
                        break
                except Exception as err:
                    err_msg = str(err)
                    sys.stderr.write(f"Gemini translation {m} error: {err_msg}\n")
                    if '429' in err_msg or 'RESOURCE_EXHAUSTED' in err_msg:
                        key_exhausted = True
                        break
                    continue
                if response and response.text:
                    break
            if response and response.text:
                break

        if response and response.text:
            res_dict = ArabicTranslationSchema.model_validate_json(response.text).model_dump()
            res_dict["nationality_ar"] = nat
            return res_dict
    except Exception as e:
        sys.stderr.write(f"Gemini translation exception: {e}\n")
    return None

def run_groq_arabic_translation(english_data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Fallback Arabic Translation using Groq API."""
    key = GROQ_API_KEY or os.getenv("GROQ_API_KEY", "").strip()
    if not key:
        return None
    models = ['llama-3.3-70b-versatile', 'llama-3.1-8b-instant', 'allam-2-7b']
    prompt = (
        f"Phonetically transliterate the person's English first_name and last_name into official Arabic script.\n"
        f"Do NOT translate nationality — set nationality_ar exactly equal to Nationality as provided.\n"
        f"First Name: {english_data.get('first_name', '')}\n"
        f"Last Name: {english_data.get('last_name', '')}\n"
        f"Nationality: {english_data.get('nationality', '')}\n"
        f"Respond ONLY with a valid JSON object formatted as:\n"
        '{"first_name_ar": "...", "last_name_ar": "...", "nationality_ar": "..."}'
    )
    for m in models:
        try:
            payload = {
                "model": m,
                "messages": [{"role": "user", "content": prompt}],
                "response_format": {"type": "json_object"},
                "temperature": 0.1
            }
            req = urllib.request.Request(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json",
                    "User-Agent": "Mozilla/5.0"
                },
                data=json.dumps(payload).encode("utf-8")
            )
            with urllib.request.urlopen(req, timeout=10) as res:
                body = json.loads(res.read().decode("utf-8"))
                content = body["choices"][0]["message"]["content"]
                res_dict = json.loads(content)
                res_dict["nationality_ar"] = english_data.get("nationality", "")
                sys.stderr.write(f"[Fallback Translation] Groq ({m}) succeeded.\n")
                return res_dict
        except Exception as e:
            sys.stderr.write(f"[Fallback Translation] Groq ({m}) error: {e}\n")
            continue
    return None

def run_cerebras_arabic_translation(english_data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Fallback Arabic Translation using Cerebras API."""
    key = CEREBRAS_API_KEY or os.getenv("CEREBRAS_API_KEY", "").strip()
    if not key:
        return None
    models = ['gpt-oss-120b', 'zai-glm-4.7', 'gemma-4-31b']
    prompt = (
        f"Phonetically transliterate the person's English first_name and last_name into official Arabic script.\n"
        f"First Name: {english_data.get('first_name', '')}\n"
        f"Last Name: {english_data.get('last_name', '')}\n"
        f"Nationality: {english_data.get('nationality', '')}\n"
        f"Respond ONLY with JSON object: {{\"first_name_ar\": \"...\", \"last_name_ar\": \"...\", \"nationality_ar\": \"...\"}}"
    )
    for m in models:
        try:
            payload = {
                "model": m,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1
            }
            req = urllib.request.Request(
                "https://api.cerebras.ai/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json",
                    "User-Agent": "Mozilla/5.0"
                },
                data=json.dumps(payload).encode("utf-8")
            )
            with urllib.request.urlopen(req, timeout=10) as res:
                body = json.loads(res.read().decode("utf-8"))
                content = body["choices"][0]["message"]["content"]
                if "```json" in content:
                    content = content.split("```json")[1].split("```")[0].strip()
                elif "```" in content:
                    content = content.split("```")[1].split("```")[0].strip()
                res_dict = json.loads(content)
                res_dict["nationality_ar"] = english_data.get("nationality", "")
                sys.stderr.write(f"[Fallback Translation] Cerebras ({m}) succeeded.\n")
                return res_dict
        except Exception as e:
            sys.stderr.write(f"[Fallback Translation] Cerebras ({m}) error: {e}\n")
            continue
    return None

def run_openrouter_arabic_translation(english_data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Fallback Arabic Translation using OpenRouter API."""
    key = OPENROUTER_API_KEY or os.getenv("OPENROUTER_API_KEY", "").strip()
    if not key:
        return None
    models = [
        'google/gemma-4-31b-it:free',
        'inclusionai/ling-3.0-tiny:free',
        'google/gemma-4-26b-a4b-it:free'
    ]
    prompt = (
        f"Phonetically transliterate English names to official Arabic script.\n"
        f"First Name: {english_data.get('first_name', '')}\n"
        f"Last Name: {english_data.get('last_name', '')}\n"
        f"Nationality: {english_data.get('nationality', '')}\n"
        f"Respond ONLY with JSON: {{\"first_name_ar\": \"...\", \"last_name_ar\": \"...\", \"nationality_ar\": \"...\"}}"
    )
    for m in models:
        try:
            payload = {
                "model": m,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1
            }
            req = urllib.request.Request(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json",
                    "HTTP-Referer": "https://localhost",
                    "User-Agent": "Mozilla/5.0"
                },
                data=json.dumps(payload).encode("utf-8")
            )
            with urllib.request.urlopen(req, timeout=10) as res:
                body = json.loads(res.read().decode("utf-8"))
                content = body["choices"][0]["message"]["content"]
                if "```json" in content:
                    content = content.split("```json")[1].split("```")[0].strip()
                elif "```" in content:
                    content = content.split("```")[1].split("```")[0].strip()
                res_dict = json.loads(content)
                res_dict["nationality_ar"] = english_data.get("nationality", "")
                sys.stderr.write(f"[Fallback Translation] OpenRouter ({m}) succeeded.\n")
                return res_dict
        except Exception as e:
            sys.stderr.write(f"[Fallback Translation] OpenRouter ({m}) error: {e}\n")
            continue
    return None

def run_arabic_translation(english_data: Dict[str, Any], api_key: Optional[str] = None) -> Dict[str, Any]:
    """Translates English passport names to Arabic (Instant Offline Rule Engine + Fast Gemini Fallback)."""
    fn = english_data.get("first_name") or english_data.get("firstName") or ""
    ln = english_data.get("last_name") or english_data.get("lastName") or ""
    nat = english_data.get("nationality") or "Pakistani"

    # Tier 1: Fast Rule & Dictionary Engine (Instant 0.001s, 100% reliable)
    fn_ar = translate_single_field_to_arabic(fn)
    ln_ar = translate_single_field_to_arabic(ln)

    if is_arabic_text(fn_ar) and (not ln or is_arabic_text(ln_ar)):
        return {
            "first_name_ar": fn_ar,
            "last_name_ar": ln_ar,
            "nationality_ar": nat
        }

    # Tier 2: Gemini API
    res = run_gemini_arabic_translation(english_data, api_key)
    if res and (is_arabic_text(res.get("first_name_ar")) or is_arabic_text(res.get("last_name_ar"))):
        return res

    # Tier 3: Return Rule Engine result
    return {
        "first_name_ar": fn_ar,
        "last_name_ar": ln_ar,
        "nationality_ar": nat
    }

# ==============================================================================
# HIGH-LEVEL BOT API FUNCTIONS
# ==============================================================================
def process_passport_image(image_bytes: bytes) -> Dict[str, Any]:
    """
    1. Call when WhatsApp user uploads passport photo.
    2. Runs Gemini OCR and checks 6-month validity rule.
    3. Stages record as 'Pending' in SQLite.
    4. Returns data + ready-to-send WhatsApp text message.
    """
    extracted_data = run_passport_ocr(image_bytes)
    if not extracted_data or not isinstance(extracted_data, dict):
        return {
            "success": False,
            "error": "Could not extract text from passport image. Please make sure the photo is clear, well-lit, un-cropped, and try sending again."
        }
    
    # 6-Month Passport Validity Check
    is_valid, validity_msg = validate_passport_validity(extracted_data.get('date_of_expiry', ''))
    if not is_valid:
        return {
            "success": False,
            "validity_error": True,
            "record": extracted_data,
            "whatsapp_message": validity_msg
        }

    saved_record = save_pending_record(extracted_data)
    
    msg = (
        "📄 *Passport Data Extracted*\n\n"
        f"• *First Name:* {extracted_data['first_name']}\n"
        f"• *Last Name:* {extracted_data['last_name']}\n"
        f"• *Passport #:* {extracted_data['passport_number']}\n"
        f"• *Nationality:* {extracted_data['nationality']}\n"
        f"• *DOB:* {extracted_data['date_of_birth']}\n"
        f"• *Issue Date:* {extracted_data.get('date_of_issue', 'N/A')}\n"
        f"• *Expiry Date:* {extracted_data['date_of_expiry']}\n\n"
        "👉 Reply *YES* to Confirm Details\n"
        "👉 Reply *NO* to Reject & Retry"
    )
    
    return {
        "success": True,
        "record": saved_record,
        "whatsapp_message": msg
    }

def confirm_and_translate_passport(passport_number: str, english_data: Optional[Dict[str, Any]] = None, phone: str = "", request_id: str = "") -> Dict[str, Any]:
    """
    1. Call when WhatsApp user replies 'YES' (Confirm).
    2. Translates details to Arabic script silently in background.
    3. Updates status to 'Confirmed' in SQLite.
    4. Automatically appends row to Master_Passports.xlsx.
    5. Returns confirmed record + simple confirmation message.
    """
    if not english_data:
        english_data = get_record(passport_number)
        if not english_data:
            return {"success": False, "error": f"No record found for passport {passport_number}"}
    else:
        save_pending_record(english_data, phone=phone, request_id=request_id)
            
    arabic_data = run_arabic_translation(english_data)
    confirmed_record = update_confirmed_record(passport_number, english_data, arabic_data, phone=phone, request_id=request_id)
    excel_path = export_confirmed_passports_to_excel(request_id)
    
    msg = "✅ *Passport Confirmed & Recorded!*"
    
    return {
        "success": True,
        "record": confirmed_record,
        "excel_file": excel_path,
        "whatsapp_message": msg
    }

def validate_ticket_date(departure_date_str: str) -> tuple[bool, str, str]:
    """
    Checks if flight departure date is strictly in the future (Travel Date > Today).
    Returns (is_valid, formatted_date, whatsapp_message)
    """
    if not departure_date_str or departure_date_str.strip().upper() in ['N/A', 'NONE', 'NOT DETECTED']:
        return False, "", "❌ *Could not detect flight departure date on ticket.*\n\nPlease make sure the photo is clear and shows your flight departure date."

    try:
        exp_date = None
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y', '%d %b %Y', '%d-%b-%Y', '%d-%b-%y']:
            try:
                exp_date = datetime.strptime(departure_date_str.strip(), fmt)
                break
            except ValueError:
                continue

        if not exp_date:
            return False, "", "❌ *Invalid date format on ticket.*\n\nPlease send a clear photo of your ticket booking."

        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        flight_date = exp_date.replace(hour=0, minute=0, second=0, microsecond=0)
        formatted_date = flight_date.strftime('%d-%b-%Y')

        if flight_date <= today:
            msg = (
                "❌ *Ticket Travel Date Invalid!*\n\n"
                f"• *Flight Departure Date:* {formatted_date}\n\n"
                "⚠️ *Rule:* Your travel departure date must be a **future date (greater than today)**.\n\n"
                "Please upload a valid ticket booking image with a future departure date to proceed."
            )
            return False, formatted_date, msg

        return True, formatted_date, f"✅ *Ticket Booking Validated!* (Departure Date: {formatted_date})"
    except Exception as e:
        sys.stderr.write(f"Ticket date validation error: {e}\n")
        return False, "", "❌ *Could not validate ticket departure date.* Please send a clear ticket photo."

def run_gemini_ticket_ocr(file_bytes: bytes, mime_type: str = "image/jpeg", api_key: Optional[str] = None) -> Optional[Dict[str, Any]]:
    keys = get_gemini_api_keys(api_key)
    try:
        from google import genai
        from google.genai import types

        if mime_type == "application/pdf" or file_bytes.startswith(b"%PDF"):
            compressed_bytes = file_bytes
            payload_mime = "application/pdf"
        else:
            compressed_bytes, payload_mime = compress_image_if_needed(file_bytes)

        prompt = """
        Extract from this flight ticket booking document or image:
        1. departure_date: Outbound flight departure date (Flight 1 / Leg 1) strictly in YYYY-MM-DD format (e.g. '27 AUG 2026' -> '2026-08-27').
        2. return_date: Return flight departure date (Flight 2 / Leg 2 return segment e.g. '17 NOV 2026' -> '2026-11-17') strictly in YYYY-MM-DD format. Look for 'Flight 2', 'Return', or second flight box.
        3. airline_name: Airline carrier name e.g. 'PIA', 'Saudi Arabian Airlines', 'Airblue'.
        4. flight_numbers: Combined flight numbers for outbound and return e.g. 'PK 731 / PK 732' or 'SV 735 / SV 734'.
        5. origin_city: Departure city or airport code e.g. 'Karachi', 'KHI', 'Lahore', 'LHE', 'Islamabad', 'ISB'.
        6. destination_city: Arrival city or airport code e.g. 'Jeddah', 'JED', 'Madinah', 'MED'.
        7. arrival_airport: Arrival airport code e.g. 'JED', 'MED'.

        CRITICAL MULTI-FLIGHT INSTRUCTION:
        - If there is a 'Flight 1' and 'Flight 2' (or return segment), return_date MUST be the departure date of Flight 2 (e.g. 17 NOV 2026 -> 2026-11-17).
        - Format all dates strictly as YYYY-MM-DD.
        """

        candidate_models = ['gemini-flash-lite-latest', 'gemini-flash-latest', 'gemini-pro-latest', 'gemini-3-flash-preview']
        response = None
        for k in keys:
            key_exhausted = False
            client = genai.Client(api_key=k)
            for m in candidate_models:
                if key_exhausted:
                    break
                for attempt in range(2):
                    try:
                        response = client.models.generate_content(
                            model=m,
                            contents=[types.Part.from_bytes(data=compressed_bytes, mime_type=payload_mime), prompt],
                            config=types.GenerateContentConfig(
                                response_mime_type="application/json",
                                response_schema=TicketSchema,
                                temperature=0.0
                            )
                        )
                        if response and response.text:
                            break
                    except Exception as err:
                        err_msg = str(err)
                        if '429' in err_msg or 'RESOURCE_EXHAUSTED' in err_msg:
                            key_exhausted = True
                            break
                        break
                if response and response.text:
                    break
            if response and response.text:
                break

        if response and response.text:
            return TicketSchema.model_validate_json(response.text).model_dump()
    except Exception as e:
        sys.stderr.write(f"Gemini ticket OCR exception: {e}\n")
    return None

def run_openrouter_ticket_ocr(file_bytes: bytes, mime_type: str = "image/jpeg") -> Optional[Dict[str, Any]]:
    key = OPENROUTER_API_KEY or os.getenv("OPENROUTER_API_KEY", "").strip()
    if not key:
        return None
    try:
        if mime_type == "application/pdf" or file_bytes.startswith(b"%PDF"):
            compressed_bytes = file_bytes
            payload_mime = "application/pdf"
        else:
            compressed_bytes, payload_mime = compress_image_if_needed(file_bytes)

        b64_data = base64.b64encode(compressed_bytes).decode("utf-8")
        data_url = f"data:{payload_mime};base64,{b64_data}"
        models = ['openrouter/free', 'google/gemma-4-31b-it:free']
        prompt = "Extract flight departure date from this ticket booking as JSON: {\"departure_date\": \"YYYY-MM-DD\"}"
        for m in models:
            try:
                payload = {
                    "model": m,
                    "messages": [{"role": "user", "content": [{"type": "text", "text": prompt}, {"type": "image_url", "image_url": {"url": data_url}}]}],
                    "temperature": 0.0
                }
                req = urllib.request.Request("https://openrouter.ai/api/v1/chat/completions", headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}, data=json.dumps(payload).encode("utf-8"))
                with urllib.request.urlopen(req, timeout=15) as res:
                    body = json.loads(res.read().decode("utf-8"))
                    content = body["choices"][0]["message"]["content"]
                    if "```json" in content: content = content.split("```json")[1].split("```")[0].strip()
                    elif "```" in content: content = content.split("```")[1].split("```")[0].strip()
                    return TicketSchema.model_validate(json.loads(content)).model_dump()
            except Exception:
                continue
    except Exception:
        pass
    return None

def run_ticket_ocr(file_bytes: bytes, is_pdf: bool = False, api_key: Optional[str] = None) -> Optional[Dict[str, Any]]:
    mime_type = "application/pdf" if (is_pdf or file_bytes.startswith(b"%PDF")) else "image/jpeg"
    res = run_gemini_ticket_ocr(file_bytes, mime_type=mime_type, api_key=api_key)
    if res:
        return res
    return run_openrouter_ticket_ocr(file_bytes, mime_type=mime_type)

def format_ticket_date_pretty(date_str: str) -> str:
    if not date_str or str(date_str).strip().upper() in ['N/A', 'NONE', 'NOT DETECTED', '']:
        return ""
    try:
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y', '%d %b %Y', '%d-%b-%Y', '%d-%b-%y']:
            try:
                dt = datetime.strptime(str(date_str).strip(), fmt)
                return dt.strftime('%d %b %Y')
            except ValueError:
                continue
    except Exception:
        pass
    return str(date_str)

AIRLINE_CODE_MAP = {
    'PAKISTAN INTERNATIONAL AIRLINES': 'PK',
    'PAKISTAN INTERNATIONAL AIRLINE': 'PK',
    'PIA': 'PK',
    'SAUDI ARABIAN AIRLINES': 'SV',
    'SAUDIA': 'SV',
    'SAUDI AIRLINES': 'SV',
    'AIRBLUE': 'PA',
    'SERENE AIR': 'ER',
    'SERENEAIR': 'ER',
    'FLY JINNAH': '9P',
    'QATAR AIRWAYS': 'QR',
    'EMIRATES': 'EK',
    'ETIHAD AIRWAYS': 'EY',
    'FLYNAS': 'XY',
    'OMAN AIR': 'WY',
    'GULF AIR': 'GF',
}

CITY_CODE_MAP = {
    'KARACHI': 'KHI',
    'LAHORE': 'LHE',
    'ISLAMABAD': 'ISB',
    'PESHAWAR': 'PEW',
    'MULTAN': 'MUX',
    'SIALKOT': 'SKT',
    'QUETTA': 'UET',
    'JEDDAH': 'JED',
    'MADINAH': 'MED',
    'MEDINA': 'MED',
}

def get_airline_short_code(airline_name: str, flight_numbers: str) -> str:
    if flight_numbers:
        match = re.search(r'([A-Z0-9]{2})\s*\d+', str(flight_numbers).upper())
        if match:
            return match.group(1)
    clean_name = str(airline_name or '').strip().upper()
    if clean_name in AIRLINE_CODE_MAP:
        return AIRLINE_CODE_MAP[clean_name]
    for key, val in AIRLINE_CODE_MAP.items():
        if key in clean_name:
            return val
    return clean_name or 'SV'

def get_city_short_code(city_name: str) -> str:
    clean_city = str(city_name or '').strip().upper()
    if clean_city in CITY_CODE_MAP:
        return CITY_CODE_MAP[clean_city]
    for key, val in CITY_CODE_MAP.items():
        if key in clean_city:
            return val
    return clean_city[:3] if clean_city else 'KHI'

def process_ticket_booking_image(file_bytes: bytes, is_pdf: bool = False) -> Dict[str, Any]:
    ticket_data = run_ticket_ocr(file_bytes, is_pdf=is_pdf)
    if not ticket_data or not isinstance(ticket_data, dict):
        return {
            "success": False,
            "error": "Could not read ticket booking details. Please ensure the departure date is clearly visible."
        }
    dep_date = ticket_data.get('departure_date', '')
    ret_date = ticket_data.get('return_date', '')
    raw_airport = str(ticket_data.get('arrival_airport') or '').upper().strip()

    arrival_airport = "UNKNOWN"
    if "MED" in raw_airport or "MADINAH" in raw_airport or "MEDINA" in raw_airport:
        arrival_airport = "MADINAH"
    elif "JED" in raw_airport or "JEDDAH" in raw_airport:
        arrival_airport = "JEDDAH"

    is_valid, formatted_date, whatsapp_msg = validate_ticket_date(dep_date)

    dep_pretty = format_ticket_date_pretty(dep_date)
    ret_pretty = format_ticket_date_pretty(ret_date) if (ret_date and ret_date.upper() != 'N/A') else ""

    if dep_pretty and ret_pretty and ret_pretty != dep_pretty:
        travel_period = f"{dep_pretty} – {ret_pretty}"
    elif dep_pretty:
        travel_period = dep_pretty
    else:
        travel_period = "Confirmed"

    airline = ticket_data.get('airline_name') or 'PIA'
    flight_nums = ticket_data.get('flight_numbers') or ''
    origin_raw = ticket_data.get('origin_city') or 'Karachi'
    dest_raw = ticket_data.get('destination_city') or (arrival_airport if arrival_airport != "UNKNOWN" else 'Jeddah')

    short_carrier = get_airline_short_code(airline, flight_nums)
    origin_code = get_city_short_code(origin_raw)
    dest_code = get_city_short_code(dest_raw)

    carrier_label = f"{short_carrier} ({flight_nums})" if flight_nums else short_carrier

    if ret_pretty:
        flight_route = f"{carrier_label} | {origin_code} → {dest_code} ({dep_pretty}) | {dest_code} → {origin_code} ({ret_pretty})"
    else:
        flight_route = f"{carrier_label} | {origin_code} → {dest_code} ({dep_pretty})"

    return {
        "success": is_valid,
        "departure_date": dep_date,
        "return_date": ret_date,
        "arrival_airport": arrival_airport,
        "formatted_date": formatted_date,
        "travel_period": travel_period,
        "airline_name": airline,
        "short_carrier": short_carrier,
        "flight_numbers": flight_nums,
        "flight_route": flight_route,
        "whatsapp_message": whatsapp_msg
    }

# ==============================================================================
# COMMAND LINE INTERFACE (For Node.js Child Process Integration)
# ==============================================================================
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No action specified. Use 'ocr <image_path>', 'ticket_ocr <image_path>', or 'confirm <passport_number> [json_data]'"}))
        sys.exit(1)

    action = sys.argv[1].lower()

    if action == "ocr":
        if len(sys.argv) < 3:
            print(json.dumps({"error": "Image file path required for OCR"}))
            sys.exit(1)
        image_path = sys.argv[2]
        if not os.path.exists(image_path):
            print(json.dumps({"error": f"Image file not found: {image_path}"}))
            sys.exit(1)
        with open(image_path, "rb") as f:
            image_bytes = f.read()
        res = process_passport_image(image_bytes)
        print(json.dumps(res))

    elif action == "ticket_ocr":
        if len(sys.argv) < 3:
            print(json.dumps({"error": "Ticket file path required for ticket_ocr"}))
            sys.exit(1)
        file_path = sys.argv[2]
        if not os.path.exists(file_path):
            print(json.dumps({"error": f"Ticket file not found: {file_path}"}))
            sys.exit(1)
        with open(file_path, "rb") as f:
            file_bytes = f.read()
        is_pdf = file_path.lower().endswith(".pdf") or file_bytes.startswith(b"%PDF")
        res = process_ticket_booking_image(file_bytes, is_pdf=is_pdf)
        print(json.dumps(res))

    elif action == "confirm":
        if len(sys.argv) < 3:
            print(json.dumps({"error": "Passport number required for confirmation"}))
            sys.exit(1)
        passport_num = sys.argv[2]
        eng_data = None
        phone = ""
        req_id = ""
        if len(sys.argv) >= 4:
            try:
                eng_data = json.loads(sys.argv[3])
            except Exception as e:
                eng_data = None
        if len(sys.argv) >= 5:
            phone = sys.argv[4]
        if len(sys.argv) >= 6:
            req_id = sys.argv[5]
        res = confirm_and_translate_passport(passport_num, eng_data, phone=phone, request_id=req_id)
        print(json.dumps(res))

    elif action == "export_excel":
        req_id = sys.argv[2] if (len(sys.argv) >= 3 and sys.argv[2] != "ALL") else None
        direct_passengers = None
        if len(sys.argv) >= 4:
            try:
                direct_passengers = json.loads(sys.argv[3])
            except Exception as e:
                direct_passengers = None
        excel_p = export_confirmed_passports_to_excel(req_id, direct_passengers=direct_passengers)
        print(json.dumps({"success": True, "excel_file": excel_p}))

    else:
        print(json.dumps({"error": f"Unknown action: {action}"}))
        sys.exit(1)
